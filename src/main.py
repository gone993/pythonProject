import time
import openpyxl
import collections as co
import datetime

from selenium import webdriver
from selenium.common.exceptions import TimeoutException, ElementClickInterceptedException
from selenium.webdriver.common.by import By
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from openpyxl import load_workbook, Workbook
import os.path

# 리뷰 크롤링
def crawling_app(url, app_name, bank_type) :
    options = webdriver.ChromeOptions()
    options.add_argument('headless')
    driverPath = "/Users/ibk/Documents/libs/chromedriver"
    # driver = webdriver.Chrome(driverPath)
    driver = webdriver.Chrome(driverPath, chrome_options=options)
    driver.get(url)

    # 관련성순 -> 최신 순서 변경
    driver.find_element_by_xpath("//span[@class='DPvwYc']").click()
    time.sleep(5)
    driver.find_element_by_xpath(
        "/html/body/div[1]/div[4]/c-wiz/div/div[2]/div/div/main/div/div[1]/div[2]/c-wiz/div[1]/div/div[2]/div[1]").click()

    SCROLL_PAUSE_TIME = 1.5
    last_height = driver.execute_script("return document.body.scrollHeight")
    while True:
        # (1) 4번의 스크롤링
        for i in range(5):
            driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
            time.sleep(SCROLL_PAUSE_TIME)
        # (2) 더보기 클릭
        try:
            WebDriverWait(driver, 10).until(
                EC.presence_of_element_located((By.XPATH, "//span[@class='RveJvd snByac']"))).click()
        except TimeoutException:
            print("timeout")
        except ElementClickInterceptedException:
            print("ElementClickInterceptedException")

        # (3) 종료 조건
        new_height = driver.execute_script("return document.body.scrollHeight")
        if new_height == last_height:
            break
        last_height = new_height

    merged_reviews = []
    reviews = driver.find_elements_by_xpath("//span[@jsname='bN97Pc']")
    for i in range(len(reviews)):
        merged_reviews.append(reviews[i].text)

    id = driver.find_elements_by_xpath("//div[@class='bAhLNe kx8XBd']/span[@class='X43Kjb']")
    dates = driver.find_elements_by_xpath("//div[@class='bAhLNe kx8XBd']/div/span[@class='p2TkOb']")
    stars = driver.find_elements_by_xpath("//span[@class='nt2C1d']/div[@class='pf5lIe']/div[@role='img']")
    likes = driver.find_elements_by_xpath("//div[@aria-label='이 리뷰가 유용하다는 평가를 받은 횟수입니다.']")

    # 타임스탬프를 찍어주기 위해 날짜 작업
    today = datetime.datetime.today().strftime("%Y%m%d")

    file_name = bank_type + '_' + today + '.xlsx'
    if (os.path.isfile(file_name)):
        wb = load_workbook(file_name)
        wb.create_sheet(title=app_name)
    else:
        wb = openpyxl.Workbook()
        sheet = wb.active
        sheet.title = app_name

    sheet = wb.get_sheet_by_name(app_name)
    sheet.append(['No', 'id', 'review', 'score', 'date'])

    print(len(merged_reviews), len(dates), len(likes), len(stars))
    res_deque = co.deque([])
    for i in range(len(dates)):
        sheet.append([i+1, id[i].text, merged_reviews[i], stars[i].get_attribute('aria-label')[10], dates[i].text])
        res_deque.append({
            'No': i + 1,
            'id': id[i].text,
            'review': merged_reviews[i],
            'score': stars[i].get_attribute('aria-label')[10],
            'date': dates[i].text
        })

    wb.save(bank_type + '_' + today + '.xlsx')

    # res_df = pd.DataFrame(res_deque)
    # res_df.to_excel('./' + 'ibk' + '_' + today + '.xlsx', index=False, encoding='utf-8-sig', sheet_name=app_name)

# 개인/기업/미니/글로벌/알림
url1 = "https://play.google.com/store/apps/details?id=com.ibk.android.ionebank&hl=ko&gl=US&showAllReviews=true"
url2 = "https://play.google.com/store/apps/details?id=com.ibk.scbs&hl=ko&gl=US&showAllReviews=true"
url3 = "https://play.google.com/store/apps/details?id=com.ibk.neobanking.silver&hl=ko&gl=US&showAllReviews=true"
url4 = "https://play.google.com/store/apps/details?id=com.ibk.neobanking.mini&hl=ko&gl=US&showAllReviews=true"
url5 = "https://play.google.com/store/apps/details?id=com.IBK.SmartPush.app&hl=ko&gl=US&showAllReviews=true"

crawling_app(url1, "user", "ibk")
crawling_app(url2, "kiup", "ibk")
crawling_app(url3, "mini", "ibk")
crawling_app(url4, "global", "ibk")
crawling_app(url5, "alrim", "ibk")

# 타행(개인)
url_kb = "https://play.google.com/store/apps/details?id=com.kbstar.kbbank&hl=ko&gl=US&showAllReviews=true"
url_sh = "https://play.google.com/store/apps/details?id=com.shinhan.sbanking&hl=ko&gl=US&showAllReviews=true"
url_wr = "https://play.google.com/store/apps/details?id=com.wooribank.smart.npib&hl=ko&gl=US&showAllReviews=true"
url_nh = "https://play.google.com/store/apps/details?id=nh.smart.banking&hl=ko&gl=US&showAllReviews=true"
url_kakao = "https://play.google.com/store/apps/details?id=com.kakaobank.channel&hl=ko&gl=US&showAllReviews=true"
url_kbank = "https://play.google.com/store/apps/details?id=com.kbankwith.smartbank&hl=ko&gl=US&showAllReviews=true"
url_toss = "https://play.google.com/store/apps/details?id=viva.republica.toss&hl=ko&gl=US&showAllReviews=true"

crawling_app(url_kb, "kb", "other")
crawling_app(url_sh, "sh", "other")
crawling_app(url_wr, "wr", "other")
crawling_app(url_nh, "nh", "other")
crawling_app(url_kakao, "kakao", "other")
crawling_app(url_kbank, "kbank", "other")
crawling_app(url_toss, "toss", "other")
